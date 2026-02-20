"""
Deep analysis of Asset Mapping Lists.xlsx
"""
import pandas as pd
import openpyxl
import re
from collections import Counter, defaultdict

FILE = r"c:\Users\nandi\Desktop\internship northladder docs\data\Asset Mapping Lists.xlsx"

print("=" * 100)
print("DEEP ANALYSIS: Asset Mapping Lists.xlsx")
print("=" * 100)

# ── 1. SHEET OVERVIEW ──────────────────────────────────────────────────────
print("\n\n" + "=" * 100)
print("SECTION 1: SHEET OVERVIEW")
print("=" * 100)

wb = openpyxl.load_workbook(FILE, read_only=True, data_only=True)
sheet_info = {}
for name in wb.sheetnames:
    ws = wb[name]
    sheet_info[name] = {
        'max_row': ws.max_row,
        'max_col': ws.max_column,
    }
    print(f"\n  Sheet: '{name}'")
    print(f"    Dimensions: {ws.max_row} rows x {ws.max_column} cols")
wb.close()

# ── 2. RAW FIRST ROWS OF EACH SHEET ───────────────────────────────────────
print("\n\n" + "=" * 100)
print("SECTION 2: RAW FIRST 8 ROWS (header=None) OF EACH SHEET")
print("=" * 100)

xls = pd.ExcelFile(FILE, engine='openpyxl')
raw_sheets = {}
for name in xls.sheet_names:
    print(f"\n--- Sheet: '{name}' ---")
    try:
        df_raw = pd.read_excel(xls, sheet_name=name, header=None, nrows=8)
        raw_sheets[name] = df_raw
        pd.set_option('display.max_columns', None)
        pd.set_option('display.width', 200)
        pd.set_option('display.max_colwidth', 60)
        print(df_raw.to_string(index=True))
    except Exception as e:
        print(f"  ERROR reading sheet: {e}")

# ── 3. INTELLIGENT RE-READ WITH PROPER HEADERS ────────────────────────────
print("\n\n" + "=" * 100)
print("SECTION 3: COLUMN NAMES (auto-detected header row)")
print("=" * 100)

sheets = {}
for name in xls.sheet_names:
    # Try header=0 first, then header=1 if first row looks like a title
    try:
        df0 = pd.read_excel(xls, sheet_name=name, header=0)
        cols0 = list(df0.columns)
        # Check if columns look like data rather than headers (e.g. all unnamed)
        unnamed_count = sum(1 for c in cols0 if 'Unnamed' in str(c))
        if unnamed_count > len(cols0) / 2:
            # Try header=1
            df1 = pd.read_excel(xls, sheet_name=name, header=1)
            cols1 = list(df1.columns)
            unnamed1 = sum(1 for c in cols1 if 'Unnamed' in str(c))
            if unnamed1 < unnamed_count:
                sheets[name] = df1
                print(f"\n  Sheet '{name}' (header=row 1): {list(df1.columns)}")
                print(f"    Shape: {df1.shape}")
                continue
        sheets[name] = df0
        print(f"\n  Sheet '{name}' (header=row 0): {cols0}")
        print(f"    Shape: {df0.shape}")
    except Exception as e:
        print(f"\n  Sheet '{name}': ERROR - {e}")

# ── 4. IDENTIFY NL MASTER SHEET ───────────────────────────────────────────
print("\n\n" + "=" * 100)
print("SECTION 4: NL MASTER LIST ANALYSIS")
print("=" * 100)

nl_sheet_name = None
nl_df = None
for name, df in sheets.items():
    name_lower = name.lower()
    if 'northladder' in name_lower or 'nl' in name_lower or 'master' in name_lower:
        nl_sheet_name = name
        nl_df = df
        break

# If not found by name, look for sheets with 'nl_product_name' or similar columns
if nl_df is None:
    for name, df in sheets.items():
        cols_lower = [str(c).lower() for c in df.columns]
        if any('nl' in c for c in cols_lower) or any('product_id' in c for c in cols_lower):
            nl_sheet_name = name
            nl_df = df
            break

if nl_df is None:
    # Just pick the largest sheet
    largest = max(sheets.items(), key=lambda x: x[1].shape[0])
    nl_sheet_name = largest[0]
    nl_df = largest[1]
    print(f"  WARNING: No obvious NL sheet found. Using largest sheet: '{nl_sheet_name}'")

print(f"\n  NL Master Sheet: '{nl_sheet_name}'")
print(f"  Shape: {nl_df.shape}")
print(f"  Columns: {list(nl_df.columns)}")
print(f"\n  First 5 rows:")
print(nl_df.head(5).to_string(index=True))
print(f"\n  Dtypes:\n{nl_df.dtypes}")

# Find key columns by fuzzy name matching
def find_col(df, keywords):
    for col in df.columns:
        col_lower = str(col).lower().strip()
        for kw in keywords:
            if kw in col_lower:
                return col
    return None

id_col = find_col(nl_df, ['asset_id', 'assetid', 'id', 'product_id', 'productid'])
name_col = find_col(nl_df, ['product_name', 'productname', 'asset_name', 'assetname', 'name', 'product'])
brand_col = find_col(nl_df, ['brand', 'make', 'manufacturer'])
cat_col = find_col(nl_df, ['category', 'cat', 'type', 'device_type'])
storage_col = find_col(nl_df, ['storage', 'capacity', 'memory', 'gb', 'rom'])

print(f"\n  Detected columns:")
print(f"    ID column:       {id_col}")
print(f"    Name column:     {name_col}")
print(f"    Brand column:    {brand_col}")
print(f"    Category column: {cat_col}")
print(f"    Storage column:  {storage_col}")

# 4a. Basic counts
print(f"\n  --- 4a. Basic Counts ---")
print(f"    Total rows: {len(nl_df)}")
if id_col:
    print(f"    Unique IDs: {nl_df[id_col].nunique()}")
    print(f"    Null IDs: {nl_df[id_col].isna().sum()}")
    dup_ids = nl_df[id_col][nl_df[id_col].duplicated(keep=False)]
    if len(dup_ids) > 0:
        print(f"    DUPLICATE IDs found: {dup_ids.nunique()} IDs appear multiple times ({len(dup_ids)} total rows)")
        for did in dup_ids.unique()[:10]:
            subset = nl_df[nl_df[id_col] == did]
            print(f"      ID={did}: {len(subset)} rows")
            if name_col:
                for _, r in subset.iterrows():
                    print(f"        -> {r.get(name_col, 'N/A')}")
if name_col:
    print(f"    Unique product names: {nl_df[name_col].nunique()}")
    print(f"    Null names: {nl_df[name_col].isna().sum()}")

# 4b. Duplicate names (same name, different IDs)
print(f"\n  --- 4b. Duplicate Names (same name, different IDs) ---")
if name_col and id_col:
    name_groups = nl_df.groupby(name_col)[id_col].apply(list).reset_index()
    dup_names = name_groups[name_groups[id_col].apply(len) > 1]
    print(f"    Names appearing with multiple IDs: {len(dup_names)}")
    for _, row in dup_names.head(15).iterrows():
        print(f"      '{row[name_col]}' -> IDs: {row[id_col]}")

# 4c. Near-duplicate names
print(f"\n  --- 4c. Near-Duplicate Names (differ by case/spaces/punctuation) ---")
if name_col:
    def normalize_name(n):
        if pd.isna(n):
            return ''
        s = str(n).lower().strip()
        s = re.sub(r'[^a-z0-9]', '', s)
        return s

    nl_df['_norm_name'] = nl_df[name_col].apply(normalize_name)
    norm_groups = nl_df.groupby('_norm_name')[name_col].apply(lambda x: list(x.unique())).reset_index()
    near_dups = norm_groups[norm_groups[name_col].apply(len) > 1]
    print(f"    Near-duplicate groups: {len(near_dups)}")
    for _, row in near_dups.head(20).iterrows():
        print(f"      Normalized='{row['_norm_name'][:50]}': variants={row[name_col]}")

# 4d. Category distribution
print(f"\n  --- 4d. Category Distribution ---")
if cat_col:
    cat_dist = nl_df[cat_col].value_counts(dropna=False)
    print(f"    Total categories: {cat_dist.shape[0]}")
    for cat, count in cat_dist.items():
        print(f"      {str(cat):40s} : {count}")
else:
    print("    No category column found.")
    # Check all columns for anything that looks like categories
    for col in nl_df.columns:
        nunique = nl_df[col].nunique()
        if 2 <= nunique <= 30 and nl_df[col].dtype == 'object':
            print(f"    Possible category column '{col}': {nunique} unique values")
            print(f"      Values: {nl_df[col].value_counts().head(10).to_dict()}")

# 4e. Custom configuration / placeholder text
print(f"\n  --- 4e. Custom Configuration / Placeholder Entries ---")
custom_patterns = ['custom config', 'placeholder', 'test', 'dummy', 'n/a', 'tbd', 'unknown', 'other', 'none']
for col in nl_df.columns:
    if nl_df[col].dtype == 'object':
        for pat in custom_patterns:
            mask = nl_df[col].astype(str).str.lower().str.contains(pat, na=False)
            count = mask.sum()
            if count > 0:
                print(f"    Column '{col}' contains '{pat}': {count} rows")
                examples = nl_df[mask][col].head(3).tolist()
                print(f"      Examples: {examples}")

# 4f. Storage encoding inconsistencies
print(f"\n  --- 4f. Storage Encoding Inconsistencies ---")
storage_patterns_found = defaultdict(list)
for col in nl_df.columns:
    if nl_df[col].dtype == 'object':
        for idx, val in nl_df[col].items():
            s = str(val)
            # Look for storage patterns
            matches = re.findall(r'\d+\s*[GT]B|\d+\.\d+\s*TB', s, re.IGNORECASE)
            for m in matches:
                storage_patterns_found[m.strip()].append((idx, col, s[:80]))

if storage_patterns_found:
    print(f"    Unique storage patterns found: {len(storage_patterns_found)}")
    # Group by numeric value
    storage_by_value = defaultdict(list)
    for pat, examples in storage_patterns_found.items():
        num = re.sub(r'[^0-9.]', '', pat)
        unit = re.sub(r'[0-9.\s]', '', pat).upper()
        storage_by_value[(num, unit)].append((pat, len(examples)))

    for (num, unit), variants in sorted(storage_by_value.items()):
        if len(variants) > 1:
            print(f"    INCONSISTENCY for {num}{unit}:")
            for pat, count in variants:
                print(f"      '{pat}': {count} occurrences")

    # Show all patterns with counts
    print(f"\n    All storage patterns:")
    for pat, examples in sorted(storage_patterns_found.items(), key=lambda x: -len(x[1])):
        print(f"      '{pat}': {len(examples)} occurrences")
else:
    print("    No storage patterns found in text columns.")

# Also check if there's a dedicated storage column
if storage_col:
    print(f"\n    Dedicated storage column '{storage_col}':")
    print(f"      Unique values: {nl_df[storage_col].nunique()}")
    print(f"      Value distribution:")
    for val, cnt in nl_df[storage_col].value_counts(dropna=False).head(30).items():
        print(f"        {str(val):20s} : {cnt}")

# 4g. Inconsistent model tokenization
print(f"\n  --- 4g. Inconsistent Model Tokenization ---")
tokenization_checks = {
    'ProMax vs Pro Max': (r'pro\s*max', [r'promax', r'pro max', r'pro  max']),
    'ZFold vs Z Fold': (r'z\s*fold', [r'zfold', r'z fold', r'z  fold']),
    'ZFlip vs Z Flip': (r'z\s*flip', [r'zflip', r'z flip']),
    'S23/S24 vs S 23/S 24': (r's\s*\d{2}', [r's\d{2}', r's \d{2}']),
    'Ultra spacing': (r'ultra', [r'ultra\b', r'\bultra']),
    'Plus spacing': (r'plus', [r'plus\b']),
    'FE spacing': (r'\bfe\b', [r'\bfe\b']),
    'iPad Pro': (r'ipad\s*pro', [r'ipadpro', r'ipad pro']),
    'iPhone spacing': (r'iphone\s*\d+', [r'iphone\d+', r'iphone \d+']),
    'MacBook': (r'mac\s*book', [r'macbook', r'mac book']),
    'AirPods': (r'air\s*pods', [r'airpods', r'air pods']),
    'Galaxy': (r'galaxy', []),
}

if name_col:
    names_lower = nl_df[name_col].astype(str).str.lower()
    for check_name, (pattern, variants) in tokenization_checks.items():
        matches = names_lower[names_lower.str.contains(pattern, na=False, regex=True)]
        if len(matches) > 0:
            # Find actual unique surface forms around the pattern
            unique_forms = set()
            for val in matches.values:
                found = re.findall(r'\b\w*' + pattern + r'\w*\b', val)
                unique_forms.update(found)
            if len(unique_forms) > 1:
                print(f"    {check_name}: {len(matches)} matches, forms: {sorted(unique_forms)[:10]}")

# 4h. Missing/empty brand
print(f"\n  --- 4h. Missing/Empty Brand ---")
if brand_col:
    null_brand = nl_df[brand_col].isna().sum()
    empty_brand = (nl_df[brand_col].astype(str).str.strip() == '').sum()
    print(f"    Null brand: {null_brand}")
    print(f"    Empty string brand: {empty_brand}")
    print(f"    Brand distribution:")
    for val, cnt in nl_df[brand_col].value_counts(dropna=False).head(30).items():
        print(f"      {str(val):30s} : {cnt}")
else:
    print("    No brand column found.")

# ── 5. MAPPING LIST SHEETS ANALYSIS ───────────────────────────────────────
print("\n\n" + "=" * 100)
print("SECTION 5: MAPPING LIST SHEETS ANALYSIS")
print("=" * 100)

mapping_sheets = {}
for name, df in sheets.items():
    if name == nl_sheet_name:
        continue
    name_lower = name.lower()
    if 'list' in name_lower or 'map' in name_lower or 'auction' in name_lower or 'asset' in name_lower:
        mapping_sheets[name] = df
    else:
        # Include all non-NL sheets
        mapping_sheets[name] = df

for name, df in mapping_sheets.items():
    print(f"\n  --- Sheet: '{name}' ---")
    print(f"    Shape: {df.shape}")
    print(f"    Columns: {list(df.columns)}")
    print(f"\n    First 5 rows:")
    print(df.head(5).to_string(index=True))

    # Find brand column
    m_brand_col = find_col(df, ['brand', 'make', 'manufacturer'])
    if m_brand_col:
        print(f"\n    Brand distribution (col='{m_brand_col}'):")
        for val, cnt in df[m_brand_col].value_counts(dropna=False).head(20).items():
            print(f"      {str(val):30s} : {cnt}")
        null_brand = df[m_brand_col].isna().sum()
        empty_brand = (df[m_brand_col].astype(str).str.strip() == '').sum()
        print(f"    Null/empty brand: {null_brand} null, {empty_brand} empty")

    # Find name/model column
    m_name_col = find_col(df, ['product_name', 'productname', 'asset_name', 'assetname', 'name', 'model', 'product', 'description', 'item'])
    if m_name_col:
        print(f"\n    Name column: '{m_name_col}' ({df[m_name_col].nunique()} unique)")
        # Variant naming examples
        variant_patterns = {
            'fold': r'fold',
            'flip': r'flip',
            'pro max': r'pro\s*max',
            'ultra': r'ultra',
            'plus': r'plus',
            'lite': r'lite',
            'mini': r'mini',
            'se': r'\bse\b',
            'fe': r'\bfe\b',
        }
        print(f"\n    Variant naming examples:")
        for vname, vpat in variant_patterns.items():
            matches = df[df[m_name_col].astype(str).str.lower().str.contains(vpat, na=False, regex=True)]
            if len(matches) > 0:
                examples = matches[m_name_col].head(3).tolist()
                print(f"      '{vname}': {len(matches)} entries, e.g. {examples}")

    # Find ID column that might reference NL
    m_id_col = find_col(df, ['nl_id', 'northladder_id', 'asset_id', 'assetid', 'matched_id', 'id', 'nl'])
    if m_id_col:
        print(f"\n    ID reference column: '{m_id_col}' ({df[m_id_col].nunique()} unique)")

# ── 6. CROSS-ANALYSIS ─────────────────────────────────────────────────────
print("\n\n" + "=" * 100)
print("SECTION 6: CROSS-ANALYSIS")
print("=" * 100)

if name_col:
    print(f"\n  Unique NL product names: {nl_df[name_col].nunique()}")

# Check mapping sheets for ID references to NL
if id_col:
    nl_ids = set(nl_df[id_col].dropna().unique())
    print(f"  NL unique IDs: {len(nl_ids)}")

    for mname, mdf in mapping_sheets.items():
        m_id_col = find_col(mdf, ['nl_id', 'northladder_id', 'asset_id', 'assetid', 'matched_id', 'nl_asset_id', 'product_id'])
        if m_id_col:
            map_ids = set(mdf[m_id_col].dropna().unique())
            missing = map_ids - nl_ids
            print(f"\n  Sheet '{mname}' -> ID col '{m_id_col}':")
            print(f"    Total IDs referenced: {len(map_ids)}")
            print(f"    IDs NOT in NL list: {len(missing)}")
            if missing:
                print(f"    Examples of missing IDs: {list(missing)[:15]}")

            # IDs that map to multiple NL assets
            if id_col and name_col:
                id_to_names = nl_df.groupby(id_col)[name_col].apply(list).to_dict()
                multi_map = {k: v for k, v in id_to_names.items() if len(v) > 1 and k in map_ids}
                if multi_map:
                    print(f"    IDs mapping to MULTIPLE NL assets: {len(multi_map)}")
                    for mid, names in list(multi_map.items())[:10]:
                        print(f"      ID={mid}: {names}")

# ── 7. ADDITIONAL: LOOK FOR ALL COLUMN VALUES THAT LOOK LIKE IDs ─────────
print("\n\n" + "=" * 100)
print("SECTION 7: ADDITIONAL - ALL NUMERIC-LIKE COLUMNS ACROSS SHEETS")
print("=" * 100)

for name, df in sheets.items():
    print(f"\n  Sheet '{name}':")
    for col in df.columns:
        if df[col].dtype in ['int64', 'float64']:
            print(f"    Numeric col '{col}': min={df[col].min()}, max={df[col].max()}, nulls={df[col].isna().sum()}")
        elif df[col].dtype == 'object':
            # Check if it looks numeric
            sample = df[col].dropna().head(5).tolist()
            print(f"    Object col '{col}': sample={sample}")

# ── 8. LOOK FOR MERGED CELLS OR MULTI-HEADER STRUCTURES ──────────────────
print("\n\n" + "=" * 100)
print("SECTION 8: MERGED CELLS / MULTI-HEADER DETECTION")
print("=" * 100)

wb = openpyxl.load_workbook(FILE, read_only=False, data_only=True)
for name in wb.sheetnames:
    ws = wb[name]
    merged = list(ws.merged_cells.ranges)
    if merged:
        print(f"\n  Sheet '{name}': {len(merged)} merged cell ranges")
        for m in merged[:20]:
            print(f"    {m}")
    else:
        print(f"\n  Sheet '{name}': No merged cells")
wb.close()

# ── 9. SPECIAL: CHECK FOR TABS/SUBTABLES WITHIN SHEETS ───────────────────
print("\n\n" + "=" * 100)
print("SECTION 9: EMPTY ROW DETECTION (possible sub-tables)")
print("=" * 100)

for name, df in sheets.items():
    all_null_rows = df.isnull().all(axis=1)
    empty_row_indices = list(all_null_rows[all_null_rows].index)
    if empty_row_indices:
        print(f"\n  Sheet '{name}': {len(empty_row_indices)} completely empty rows at indices: {empty_row_indices[:20]}")
    else:
        print(f"\n  Sheet '{name}': No completely empty rows")

# Cleanup temp column
if '_norm_name' in nl_df.columns:
    nl_df.drop(columns=['_norm_name'], inplace=True)

print("\n\n" + "=" * 100)
print("ANALYSIS COMPLETE")
print("=" * 100)
